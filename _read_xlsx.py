import openpyxl
import json

wb = openpyxl.load_workbook(r'C:\Users\HP\OneDrive\Desktop\smart_farm\smart_farm\All_Crops_Data.xlsx', data_only=True)

ws = wb[wb.sheetnames[0]]
print(f'Total Rows: {ws.max_row}, Total Cols: {ws.max_column}')
print()

for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False), 1):
    vals = [str(cell.value).strip() if cell.value is not None else '' for cell in row]
    print(f'Row {row_idx}: {json.dumps(vals)}')
